from transaction import Transaction
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat

NumberFormat.FORMAT_CURRENCY_EUR = "[$" + u"\u20AC" + " ]#,##0.00_-"

COL_DST_CHARGEDDATE = 0
COL_DST_ORDERNO = 1
COL_DST_CURRENCY = 2
COL_DST_COUNTRY = 3
COL_DST_SALESINCLVAT = 4
COL_DST_SALESEXCLVAT = 5
COL_DST_SERVICEFEE = 6
COL_DST_BALANCE = 7
COL_DST_VAT = 8
COL_DST_PAYOUTDATE = 9

header = {}
header[COL_DST_CHARGEDDATE] = 'Charged date'
header[COL_DST_ORDERNO] = 'Merchant order number'
header[COL_DST_CURRENCY] = 'Currency'
header[COL_DST_COUNTRY] = 'Country'
header[COL_DST_SALESINCLVAT] = 'Sales incl. VAT'
header[COL_DST_SALESEXCLVAT] = 'Sales excl. VAT'
header[COL_DST_SERVICEFEE] = 'Google service fee in EUR'
header[COL_DST_BALANCE] = 'Balance in EUR'
header[COL_DST_VAT] = 'VAT'
header[COL_DST_PAYOUTDATE] = 'Payout date'

sumcols = [COL_DST_SALESINCLVAT, COL_DST_SALESEXCLVAT, COL_DST_SERVICEFEE, COL_DST_BALANCE, COL_DST_VAT]

def saveasworkbook(tadic, outputfilename):
	
	workbook = Workbook()
	sheet = workbook.worksheets[0]

	row = 0 # cursor for the output row
	totalrows = [] # list of totals rows
	for currency, talist in tadic.iteritems():
		row += createheader(sheet, row)
		grouprows = creategroup(sheet, row, currency, talist)
		row += grouprows
		row += createsubtotals(sheet, row, grouprows, currency)
		totalrows.append(row)
		row += 1 # separator row
	
	# set column widths manually
	sheet.column_dimensions[get_column_letter(COL_DST_CHARGEDDATE+1)].width = 12
	sheet.column_dimensions[get_column_letter(COL_DST_ORDERNO+1)].width = 20
	sheet.column_dimensions[get_column_letter(COL_DST_CURRENCY+1)].width = 9
	sheet.column_dimensions[get_column_letter(COL_DST_COUNTRY+1)].width = 9
	sheet.column_dimensions[get_column_letter(COL_DST_SALESINCLVAT+1)].width = 13
	sheet.column_dimensions[get_column_letter(COL_DST_SALESEXCLVAT+1)].width = 13
	sheet.column_dimensions[get_column_letter(COL_DST_SERVICEFEE+1)].width = 21
	sheet.column_dimensions[get_column_letter(COL_DST_BALANCE+1)].width = 13
	sheet.column_dimensions[get_column_letter(COL_DST_VAT+1)].width = 7
	sheet.column_dimensions[get_column_letter(COL_DST_PAYOUTDATE+1)].width = 12

	# create EUR totals
	sheet.cell(row=row, column=COL_DST_SALESEXCLVAT).value = 'Total EUR'
	sheet.cell(row=row, column=COL_DST_SALESEXCLVAT).style.font.bold = True
	feesumcells = []
	feecolletter = get_column_letter(COL_DST_SERVICEFEE+1)
	balancesumcells = []
	balancecolletter = get_column_letter(COL_DST_BALANCE+1)
	vatsumcells = []
	vatcolletter = get_column_letter(COL_DST_VAT+1)
	for totalrow in totalrows:
		feesumcells.append(feecolletter + str(totalrow))
		balancesumcells.append(balancecolletter + str(totalrow))
		vatsumcells.append(vatcolletter + str(totalrow))
	sheet.cell(row=row, column=COL_DST_SERVICEFEE).value = '=SUM(' + ','.join(cell for cell in feesumcells) + ')' 
	sheet.cell(row=row, column=COL_DST_SERVICEFEE).style.font.bold = True

	sheet.cell(row=row, column=COL_DST_BALANCE).value = '=SUM(' + ','.join(cell for cell in balancesumcells) + ')' 
	sheet.cell(row=row, column=COL_DST_BALANCE).style.font.bold = True

	sheet.cell(row=row, column=COL_DST_VAT).value = '=SUM(' + ','.join(cell for cell in vatsumcells) + ')' 
	sheet.cell(row=row, column=COL_DST_VAT).style.font.bold = True

	# save the workbook
	workbook.save(outputfilename)

# writes the transaction values and returns the number of rows in the group
def creategroup(sheet, row, currency, talist):
	for ta in talist:
		sheet.cell(row=row, column=COL_DST_CHARGEDDATE).value = ta.chargeddate

		sheet.cell(row=row, column=COL_DST_ORDERNO).style.number_format.format_code = NumberFormat.FORMAT_TEXT
		sheet.cell(row=row, column=COL_DST_ORDERNO).value = ta.orderno
		
		sheet.cell(row=row, column=COL_DST_CURRENCY).value = ta.currency

		sheet.cell(row=row, column=COL_DST_COUNTRY).style.number_format.format_code = NumberFormat.FORMAT_TEXT
		sheet.cell(row=row, column=COL_DST_COUNTRY).value = ta.country

		sheet.cell(row=row, column=COL_DST_SALESINCLVAT).value = ta.salesinclvat
		sheet.cell(row=row, column=COL_DST_SALESEXCLVAT).value = ta.salesexclvat
		
		sheet.cell(row=row, column=COL_DST_SERVICEFEE).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR
		sheet.cell(row=row, column=COL_DST_SERVICEFEE).value = ta.servicefee()

		sheet.cell(row=row, column=COL_DST_BALANCE).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR
		sheet.cell(row=row, column=COL_DST_BALANCE).value = ta.balance
		
		sheet.cell(row=row, column=COL_DST_VAT).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR
		sheet.cell(row=row, column=COL_DST_VAT).value = ta.vat

		sheet.cell(row=row, column=COL_DST_PAYOUTDATE).value = ta.payoutdate
		row += 1

	return len(talist)

# writes the subtotal formulas
def createsubtotals(sheet, row, grouprows, currency):
	sheet.cell(row=row, column=COL_DST_CURRENCY).value = 'Total ' + currency
	sheet.cell(row=row, column=COL_DST_CURRENCY).style.font.bold = True	
	for sumcol in sumcols:
		colletter = get_column_letter(sumcol+1)
		cell = sheet.cell(row=row, column=sumcol)
		cell.value = '=SUM(' + colletter + str(row-grouprows+1) + ':' + colletter + str(row) + ')'
		cell.style.font.bold = True

	sheet.cell(row=row, column=COL_DST_SERVICEFEE).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR
	sheet.cell(row=row, column=COL_DST_BALANCE).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR
	sheet.cell(row=row, column=COL_DST_VAT).style.number_format.format_code = NumberFormat.FORMAT_CURRENCY_EUR

	return 1

# writes the header captions
def createheader(sheet, row):
	for column, text in header.iteritems():
		sheet.cell(row=row, column=column).value = text
	return 1

